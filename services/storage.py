# services/storage.py
import openpyxl
import os
from datetime import datetime
from config import EXCEL_FILE

def load_workbook_and_sheet():
    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook.active
        print(f"📁 Planilha carregada de: {os.path.abspath(EXCEL_FILE)}")
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Data", "Descrição", "Valor", "Tipo", "Categoria"])
        workbook.save(EXCEL_FILE)
        print(f"📁 Nova planilha criada em: {os.path.abspath(EXCEL_FILE)}")
    return workbook, sheet

def add_record(workbook, sheet, linha, valor, tipo, categoria):
    print(f"🔄 TENTANDO SALVAR:")
    print(f"   📝 Descrição: {linha}")
    print(f"   💰 Valor: {valor}")
    print(f"   📂 Categoria: {categoria}")
    print(f"   📁 Arquivo: {os.path.abspath(EXCEL_FILE)}")
    
    try:
        # Adicionar linha
        nova_linha = [datetime.now().strftime("%d/%m/%Y"), linha, valor, tipo, categoria]
        sheet.append(nova_linha)
        print(f"✅ Linha adicionada à memória: {nova_linha}")
        
        # Salvar arquivo
        workbook.save(EXCEL_FILE)
        print(f"💾 Arquivo salvo com sucesso!")
        
        # Verificar se realmente salvou
        total_linhas = sheet.max_row
        print(f"📊 Total de linhas na planilha: {total_linhas}")
        
        # Ler última linha para confirmar
        ultima_linha = list(sheet.iter_rows(min_row=total_linhas, max_row=total_linhas, values_only=True))[0]
        print(f"📋 Última linha salva: {ultima_linha}")
        
    except PermissionError as e:
        print(f"❌ ERRO DE PERMISSÃO: {e}")
        print("   💡 Feche o arquivo Excel se estiver aberto!")
    except Exception as e:
        print(f"❌ ERRO INESPERADO: {e}")
        print(f"   📍 Tipo do erro: {type(e)}")
        raise e


def get_all_records(sheet):
    return list(sheet.iter_rows(min_row=2, values_only=True))

def update_record(workbook, sheet, idx, nova_categoria):
    row = list(sheet.iter_rows(min_row=2, max_row=idx+1, values_only=False))[-1]
    row[4].value = nova_categoria
    workbook.save(EXCEL_FILE)

def delete_records(workbook, sheet, indices):
    for idx in sorted(indices, reverse=True):
        sheet.delete_rows(idx+1)
    workbook.save(EXCEL_FILE)
